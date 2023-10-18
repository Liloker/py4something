from openpyxl import Workbook
from openpyxl import load_workbook
import threading

# 在全局范围内创建锁对象
excel_lock = threading.Lock()


def write_to_excel(target, status_code, response_content):
    # 打开现有的 Excel 表格
    workbook = load_workbook('通达OA-moare-RCE-Response-Data.xlsx')

    # 选择第一个工作表
    sheet = workbook.active

    # 在工作表中追加状态码和响应内容
    with excel_lock:  # 在写入前获取锁
        sheet.append([target, status_code, response_content])

    # 保存 Excel 表格
    workbook.save('通达OA-moare-RCE-Response-Data.xlsx')

# 在需要写入Excel的地方调用write_to_excel函数，传入目标、状态码和响应内容
# write_to_excel(target, response.status_code, formatted_response)


    # # 创建一个新的Excel工作簿
    # workbook = Workbook()
    #
    # # 创建一个工作表
    # sheet = workbook.active
    # sheet.title = "第一页"
    #
    # # 将响应状态码写入表格中
    # sheet.cell(row=1, column=1, value="响应状态码")
    # sheet.cell(row=2, column=1, value=response.status_code)
    #
    # # 将响应内容写入表格中
    # sheet.cell(row=1, column=2, value="响应内容")
    # sheet.cell(row=2, column=2, value=formatted_response)